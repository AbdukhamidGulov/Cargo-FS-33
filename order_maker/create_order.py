from asyncio import get_running_loop, sleep, TimeoutError
from os import path, makedirs
from shutil import rmtree
from logging import getLogger
from typing import List, Dict
from uuid import uuid4

from aiogram import Router, F, Bot
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, CallbackQuery, FSInputFile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image as PilImage

from keyboards.user_keyboards import cancel_keyboard, main_keyboard, get_order_keyboard

order_router = Router()
logger = getLogger(__name__)

TEMP_FOLDER = "temp_orders"
IMAGE_SIZE = (120, 120)


class OrderItemsStates(StatesGroup):
    waiting_for_photo = State()
    waiting_for_quantity = State()
    waiting_for_track_code = State()
    waiting_for_link = State()
    confirm_next_step = State()


def generate_excel_sync(
    filename: str,
    form_title: str,
    client_data: Dict,
    items: List[Dict],
    temp_paths: List[str]
) -> None:
    logger.info(
        "Начато формирование Excel. file=%s form_title=%s client_id=%s items_count=%s",
        filename,
        form_title,
        client_data.get("id"),
        len(items)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = form_title

    headers_info = [
        ("B1", "ID Клиента:", client_data.get("id")),
        ("B2", "Имя клиента:", client_data.get("name")),
        ("B3", "Email:", client_data.get("email"))
    ]

    for cell, title, value in headers_info:
        ws[cell] = title
        ws[cell].font = Font(bold=True)
        ws[cell.replace("B", "C")] = value

    table_headers = ["№", "Фото", "Количество", "Трек-номер", "Ссылка/Описание"]
    start_row = 7

    ws.append([])
    ws.append([])
    ws.append(table_headers)

    col_widths = {"A": 5, "B": 18, "C": 15, "D": 20, "E": 40}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for cell in ws[start_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))

    for i, item in enumerate(items, start=1):
        row = start_row + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")

        photo_path = item.get("local_path")

        if photo_path and path.exists(photo_path):
            try:
                resized_path = photo_path.replace("raw_", "res_")

                with PilImage.open(photo_path) as img:
                    img.thumbnail(IMAGE_SIZE)
                    img.save(resized_path)

                temp_paths.append(resized_path)

                img_obj = ExcelImage(resized_path)
                ws.add_image(img_obj, f"B{row}")
                ws.row_dimensions[row].height = 90

            except Exception as e:
                logger.error("Ошибка обработки изображения для Excel: %s", e, exc_info=True)
                ws.cell(row=row, column=2, value="Error")
        else:
            ws.cell(row=row, column=2, value="Нет фото")

        ws.cell(row=row, column=3, value=item["quantity"]).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        ws.cell(row=row, column=4, value=item["track"]).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        ws.cell(row=row, column=5, value=item["link"]).alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

    wb.save(filename)
    logger.info("Excel успешно сформирован: %s", filename)


async def safe_download(bot: Bot, file_path: str, destination: str, retries: int = 3) -> None:
    for attempt in range(1, retries + 1):
        try:
            logger.info(
                "Попытка скачивания файла. file_path=%s destination=%s attempt=%s/%s",
                file_path,
                destination,
                attempt,
                retries
            )
            await bot.download_file(file_path, destination=destination)
            logger.info("Файл успешно скачан: %s", destination)
            return
        except TimeoutError:
            logger.warning(
                "Timeout при скачивании файла. file_path=%s destination=%s attempt=%s/%s",
                file_path,
                destination,
                attempt,
                retries
            )
            if attempt == retries:
                raise
            await sleep(2)


async def start_item_collection(message: Message, state: FSMContext):
    logger.info("Старт заполнения бланка. tg_id=%s", message.from_user.id)

    await message.answer(
        "✅ Данные подтверждены.\n\n"
        "📦 <b>Товар №1</b>\n📸 <b>Отправьте фото товара</b>:",
        reply_markup=cancel_keyboard
    )
    await state.set_state(OrderItemsStates.waiting_for_photo)


async def cancel_order_handler(message: Message, state: FSMContext):
    logger.info("Пользователь отменил заполнение бланка. tg_id=%s", message.from_user.id)

    await state.clear()
    await message.answer("Действие отменено.", reply_markup=main_keyboard)


@order_router.message(OrderItemsStates.waiting_for_photo)
async def process_photo(message: Message, state: FSMContext):
    if message.text and message.text.lower() == "отмена":
        await cancel_order_handler(message, state)
        return

    if not message.photo:
        await message.answer("Пожалуйста, отправьте <b>сжатое фото</b>, а не файл.")
        return

    await state.update_data(current_photo=message.photo[-1].file_id)

    logger.info(
        "Получено фото товара. tg_id=%s file_id=%s",
        message.from_user.id,
        message.photo[-1].file_id
    )

    await message.answer("🔢 <b>Введите количество</b>:")
    await state.set_state(OrderItemsStates.waiting_for_quantity)


@order_router.message(OrderItemsStates.waiting_for_quantity, F.text)
async def process_quantity(message: Message, state: FSMContext):
    if message.text.lower() == "отмена":
        await cancel_order_handler(message, state)
        return

    await state.update_data(current_quantity=message.text)

    logger.info(
        "Получено количество товара. tg_id=%s quantity=%s",
        message.from_user.id,
        message.text
    )

    await message.answer("🚚 <b>Введите трек-номер</b>:")
    await state.set_state(OrderItemsStates.waiting_for_track_code)


@order_router.message(OrderItemsStates.waiting_for_track_code, F.text)
async def process_track(message: Message, state: FSMContext):
    if message.text.lower() == "отмена":
        await cancel_order_handler(message, state)
        return

    await state.update_data(current_track=message.text)

    logger.info(
        "Получен трек-номер. tg_id=%s track=%s",
        message.from_user.id,
        message.text
    )

    await message.answer("🔗 <b>Ссылка на товар</b> или описание:")
    await state.set_state(OrderItemsStates.waiting_for_link)


@order_router.message(OrderItemsStates.waiting_for_link, F.text)
async def process_link(message: Message, state: FSMContext):
    if message.text.lower() == "отмена":
        await cancel_order_handler(message, state)
        return

    data = await state.get_data()
    items = data.get("items", [])

    new_item = {
        "photo_id": data.get("current_photo"),
        "quantity": data.get("current_quantity"),
        "track": data.get("current_track"),
        "link": message.text
    }

    items.append(new_item)
    await state.update_data(items=items)

    logger.info(
        "Товар добавлен в бланк. tg_id=%s items_count=%s",
        message.from_user.id,
        len(items)
    )

    await message.answer(
        f"✅ Товар №{len(items)} сохранен!\nЧто дальше?",
        reply_markup=get_order_keyboard()
    )
    await state.set_state(OrderItemsStates.confirm_next_step)


@order_router.callback_query(OrderItemsStates.confirm_next_step, F.data == "order_add_next")
async def add_next_item(callback: CallbackQuery, state: FSMContext):
    try:
        await callback.answer()
    except Exception:
        pass

    data = await state.get_data()
    next_num = len(data.get("items", [])) + 1

    logger.info(
        "Пользователь выбрал добавление следующего товара. tg_id=%s next_item_num=%s",
        callback.from_user.id,
        next_num
    )

    if isinstance(callback.message, Message):
        try:
            await callback.message.delete()
        except Exception as e:
            logger.warning("Не удалось удалить сообщение после order_add_next: %s", e)

        await callback.message.answer(
            f"📦 <b>Товар №{next_num}</b>\n📸 <b>Отправьте фото:</b>",
            reply_markup=cancel_keyboard
        )

    await state.set_state(OrderItemsStates.waiting_for_photo)


@order_router.callback_query(OrderItemsStates.confirm_next_step, F.data == "order_finish")
async def finish_order(callback: CallbackQuery, state: FSMContext, bot: Bot):
    try:
        await callback.answer()
    except Exception:
        pass

    data = await state.get_data()
    items = data.get("items", [])

    logger.info(
        "Начато завершение бланка. tg_id=%s items_count=%s",
        callback.from_user.id,
        len(items)
    )

    if not items:
        logger.info("Попытка завершить пустой бланк. tg_id=%s", callback.from_user.id)
        await callback.message.answer("Список пуст.", reply_markup=main_keyboard)
        await state.clear()
        return

    await callback.message.edit_text("⏳ Подготавливаю файлы...")

    client_id_str = data.get("client_excel_id", str(callback.from_user.id))
    order_uid = uuid4().hex[:8]
    order_temp_dir = path.join(TEMP_FOLDER, f"{client_id_str}_{order_uid}")

    makedirs(order_temp_dir, exist_ok=True)

    logger.info(
        "Создана временная папка заказа. tg_id=%s order_temp_dir=%s",
        callback.from_user.id,
        order_temp_dir
    )

    excel_name = path.join(order_temp_dir, f"order_{client_id_str}.xlsx")
    temp_files_to_clean = []

    try:
        for idx, item in enumerate(items, 1):
            if not item.get("photo_id"):
                logger.info(
                    "У товара нет photo_id, скачивание пропущено. tg_id=%s item_num=%s",
                    callback.from_user.id,
                    idx
                )
                continue

            await callback.message.edit_text(f"⏳ Скачиваю фото {idx}/{len(items)}")

            file_path = path.join(order_temp_dir, f"raw_{idx}.jpg")
            temp_files_to_clean.append(file_path)
            item["local_path"] = file_path

            logger.info(
                "Запрошена информация о файле Telegram. tg_id=%s item_num=%s",
                callback.from_user.id,
                idx
            )

            file_info = await bot.get_file(item["photo_id"])

            await safe_download(
                bot=bot,
                file_path=file_info.file_path,
                destination=file_path
            )

        loop = get_running_loop()

        client_data = {
            "id": client_id_str,
            "name": data.get("client_name", "Unknown"),
            "email": data.get("client_email", "-")
        }

        await callback.message.edit_text("⏳ Формирую Excel файл...")

        await loop.run_in_executor(
            None,
            generate_excel_sync,
            excel_name,
            data.get("form_title", "Заказ"),
            client_data,
            items,
            temp_files_to_clean
        )

        display_name = f"{data.get('form_title', 'Order')}_{client_id_str}.xlsx"
        file_doc = FSInputFile(excel_name, filename=display_name)

        logger.info(
            "Отправка Excel пользователю. tg_id=%s file=%s",
            callback.from_user.id,
            excel_name
        )

        await callback.message.answer_document(
            file_doc,
            caption=f"✅ <b>Готово!</b>\nКлиент: {client_id_str}\nПозиций: {len(items)}",
            reply_markup=main_keyboard
        )

        logger.info("Excel успешно отправлен пользователю. tg_id=%s", callback.from_user.id)

    except TimeoutError:
        logger.error(
            "Timeout при скачивании файлов заказа. tg_id=%s",
            callback.from_user.id,
            exc_info=True
        )
        await callback.message.answer("⚠️ Telegram долго отвечает. Попробуйте снова.")

    except Exception as e:
        logger.error(
            "Ошибка создания заказа. tg_id=%s error=%s",
            callback.from_user.id,
            e,
            exc_info=True
        )
        await callback.message.answer("Произошла ошибка при создании файла.")

    finally:
        await sleep(1)

        try:
            if path.exists(order_temp_dir):
                rmtree(order_temp_dir)
                logger.info(
                    "Временная папка заказа удалена. tg_id=%s order_temp_dir=%s",
                    callback.from_user.id,
                    order_temp_dir
                )
        except Exception as e:
            logger.error("Ошибка очистки временной папки: %s", e, exc_info=True)

        await state.clear()
        logger.info("FSM состояния заказа очищены. tg_id=%s", callback.from_user.id)
